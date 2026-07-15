"""Blotter safety check: reconcile four end-of-day trade files against each other.

Sources:
    1. EMSX grid export (.xlsx)         -- ticker + SEDOL, fills, avg price
    2. Vantage Blotter EOD (.xlsx/.csv) -- S/P/C rows, bl/sl/ss/cs type codes
    3. PCMBlotter (.csv)                -- Cascade-style, accounting negatives
    4. Goldman tradefile (.csv)         -- pipe-delimited ORDER/ALLOCAT, SEDOL-keyed

Every trade is normalized to (ticker, side, quantity, price) where side is one of
BUY (bl), SELL (sl), SHORT (ss), COVER (cs). Only EMSX rows with Status "Filled"
are loaded. Buckets are aggregated per (ticker, side): summed quantity and
quantity-weighted average price.

Workflow (all comparisons bidirectional; a bucket present on one side but not the
other is a discrepancy):
  Step 1: EMSX filtered to brokers GSPT/PREX/GSOP  vs  Goldman tradefile.
  Step 2: EMSX excluding broker TDAI               vs  PCMBlotter.
  Step 3: PCM filtered to brokers GSPT/PREX/GSOP   vs  Goldman tradefile.
  Step 4: Blotter EOD adjudication. Discrepancies from steps 1-3 that are pure
          side-classification disagreements (net signed quantity per ticker still
          matches) are resolved to INFO when the Blotter EOD corroborates one
          side. Missing/extra trades that change net quantity remain errors.

Internal fund-to-fund crosses (broker CROS) never route through EMSX; they are
split out of the Blotter/PCM frames, checked to net flat per ticker, and their
volume compared between the two sources.

SEDOL -> ticker resolution order for the tradefile:
    1. EMSX file itself (has both identifiers)
    2. Backtester security master (local parquet, no network)
    3. xbbg / Bloomberg terminal (if installed and running)

Usage (CLI):
    python blotter_check.py --emsx grid.xlsx --blotter BlotterEOD.xlsx \
        --pcm "PCMBlotter 2026-07-10.csv" --tradefile tradefile.nine82.XXXX.csv \
        [--out report.csv] [--qty-tol 0] [--price-tol 0.01]

Web frontend: python eod.py (see templates/index.html).
"""

import argparse
import csv
import os
import sys
from datetime import datetime
from io import BytesIO, StringIO

import pandas as pd
from openpyxl import load_workbook

BACKTESTER_REPO = os.path.join(os.path.expanduser("~"), "GitHub", "Backtester")

SIDES = ("BUY", "SELL", "SHORT", "COVER")
SIDE_SIGN = {"BUY": 1, "COVER": 1, "SELL": -1, "SHORT": -1}

EMSX_SIDE_MAP = {
    "buy": "BUY",
    "buy to open": "BUY",
    "sell": "SELL",
    "sell to close": "SELL",
    "sell short": "SHORT",
    "short": "SHORT",
    "short sell": "SHORT",
    "buy to close": "COVER",
    "buy to cover": "COVER",
}
BLOTTER_SIDE_MAP = {"bl": "BUY", "sl": "SELL", "ss": "SHORT", "cs": "COVER"}
PCM_SIDE_MAP = EMSX_SIDE_MAP
TRADEFILE_SIDE_MAP = {"B": "BUY", "S": "SELL", "SS": "SHORT", "BC": "COVER"}

TRADEFILE_ACCOUNT_NAMES = {"065465783": "gs-top", "065448128": "gs-wsf"}

# Goldman-custodied broker codes (EMSX column G / PCM broker column).
DEFAULT_GOLDMAN_BROKERS = {"GSPT", "GSOP", "PREX"}
# EMSX brokers excluded from the PCM comparison (step 2).
EMSX_PCM_EXCLUDED_BROKERS = {"TDAI"}
# Internal fund-to-fund crosses; never routed through EMSX.
CROSS_BROKER = "CROS"

EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb"}


class Findings:
    """Accumulates flagged differences."""

    def __init__(self):
        self.rows = []

    def add(self, severity, source, check, ticker="", side="", left_value="", right_value="",
            detail="", step="Load"):
        self.rows.append({
            "severity": severity,
            "step": step,
            "source": source,
            "check": check,
            "ticker": ticker,
            "side": side,
            "left_value": left_value,
            "right_value": right_value,
            "detail": detail,
        })

    def to_frame(self):
        df = pd.DataFrame(self.rows, columns=[
            "severity", "step", "source", "check", "ticker", "side",
            "left_value", "right_value", "detail",
        ])
        order = {"ERROR": 0, "WARN": 1, "RESOLVED": 2, "INFO": 3}
        return df.sort_values(
            by=["severity", "step", "source", "ticker"],
            key=lambda s: s.map(order) if s.name == "severity" else s,
        ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _num(value):
    """Scalar to float, tolerating thousands separators from Excel pastes."""
    return pd.to_numeric(str(value).replace(",", "").strip(), errors="coerce")


def _num_series(series):
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    )


def _table_from_text(text):
    """Pasted grid (tab-separated from Excel, or raw CSV) -> headerless DataFrame."""
    sep = "\t" if "\t" in text else ","
    return pd.read_csv(
        StringIO(text), sep=sep, header=None, dtype=object,
        engine="python", on_bad_lines="skip",
    )


def _table_from_excel_bytes(data):
    return pd.read_excel(BytesIO(data), header=None)


def _find_header_row(df, needle):
    """Index of the first row containing a cell equal to `needle`, or None."""
    for i in range(min(len(df), 40)):
        if any(str(c).strip() == needle for c in df.iloc[i]):
            return i
    return None


def _named_table(df, needle):
    """Split a headerless table at its header row (identified by `needle`)."""
    hdr = _find_header_row(df, needle)
    if hdr is None:
        raise ValueError(f"Could not find a header row containing {needle!r}. "
                         "Paste or upload the file including its header row.")
    columns = [str(c).strip() if pd.notna(c) else "" for c in df.iloc[hdr]]
    body = df.iloc[hdr + 1:].reset_index(drop=True)
    body.columns = columns
    return body


# ---------------------------------------------------------------------------
# Loaders: each returns a DataFrame with columns
#   ticker, side, qty (abs, int), price (float), broker, plus source extras
# ---------------------------------------------------------------------------

def _emsx_from_table(raw, findings):
    df = _named_table(raw, "Security")
    df = df[df["Security"].notna() & (df["Security"].astype(str).str.strip() != "")].copy()
    df = df[
        df["Status"].astype(str).str.strip().str.casefold().eq("filled")
    ].copy()

    def map_side(value):
        side = EMSX_SIDE_MAP.get(str(value).strip().lower())
        if side is None:
            findings.add("ERROR", "EMSX", "unknown_side", detail=f"Unmapped EMSX side {value!r}")
        return side

    out = pd.DataFrame({
        "ticker": df["Security"].astype(str).str.strip().str.upper(),
        "sedol": df["SEDOL"].astype(str).str.strip().str.upper(),
        "side": df["Side"].map(map_side),
        "qty": _num_series(df["FillQty"]).fillna(0).astype(int),
        "price": _num_series(df["AvgPx"]),
        "broker": df["Def Brkr Code"].astype(str).str.strip().str.upper(),
        "status": df["Status"].astype(str).str.strip(),
        "ordered_qty": _num_series(df["Qty"]).fillna(0).astype(int),
    })

    partial = out[out["qty"] != out["ordered_qty"]]
    for _, r in partial.iterrows():
        findings.add(
            "WARN", "EMSX", "partial_or_unfilled", r["ticker"], r["side"] or "",
            left_value=f"ordered {r['ordered_qty']}",
            right_value=f"filled {r['qty']} ({r['status']})",
            detail="EMSX order not fully filled",
        )

    return out[out["qty"] != 0].reset_index(drop=True)


def load_emsx(path, findings):
    """EMSX grid export from a path. Non-standard xlsx internals, so use openpyxl."""
    with open(path, "rb") as f:
        return load_emsx_bytes(f.read(), str(path), findings)


def load_emsx_bytes(data, filename, findings):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in EXCEL_EXTS or not ext:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        raw = pd.DataFrame(list(ws.iter_rows(values_only=True)))
        wb.close()
    else:
        raw = _table_from_text(data.decode("utf-8", errors="replace"))
    return _emsx_from_table(raw, findings)


def load_emsx_text(text, findings):
    return _emsx_from_table(_table_from_text(text), findings)


def _blotter_from_df(df, findings):
    """Vantage BlotterEOD as a headerless table. Trade rows have S/P/C in col 0.

    S = single trade, P = parent (total across funds), C = child allocation.
    We take S and P rows; C rows duplicate their parent's total.
    Fill qty is col 8 (current fill data), falling back to col 6 (hand entered);
    fill price is col 9, falling back to col 7.
    """
    df = df[df.iloc[:, 0].isin(["S", "P", "C"])].reset_index(drop=True)

    rows = []
    unfilled = []
    for _, r in df.iterrows():
        if r.iloc[0] not in ("S", "P"):
            continue
        code = str(r.iloc[2]).strip().lower()
        side = BLOTTER_SIDE_MAP.get(code)
        ticker = str(r.iloc[3]).strip().upper()
        if side is None:
            findings.add("ERROR", "Blotter", "unknown_side", ticker,
                         detail=f"Unmapped blotter trade type {code!r}")
            continue
        broker = str(r.iloc[11]).strip().upper()
        fill_qty = _num(r.iloc[8])
        fill_px = _num(r.iloc[9])
        if not fill_qty or pd.isna(fill_qty):
            fill_qty = _num(r.iloc[6])
            fill_px = _num(r.iloc[7])
        if not fill_qty or pd.isna(fill_qty):
            if broker == CROSS_BROKER:
                # Internal crosses never show fills in the blotter; use ordered shares.
                fill_qty = _num(r.iloc[4])
                fill_px = None
                if not fill_qty or pd.isna(fill_qty):
                    continue
            else:
                unfilled.append(f"{ticker} {side} {int(_num(r.iloc[4]) or 0)} ({broker})")
                continue
        rows.append({
            "ticker": ticker,
            "side": side,
            "qty": int(fill_qty),
            "price": float(fill_px) if fill_px is not None and pd.notna(fill_px) else None,
            "broker": broker,
            "fund": str(r.iloc[18]).strip(),
        })

    if unfilled:
        findings.add("WARN", "Blotter", "unfilled_orders",
                     detail=f"{len(unfilled)} blotter order(s) with zero fill: " + "; ".join(unfilled))

    return pd.DataFrame(rows, columns=["ticker", "side", "qty", "price", "broker", "fund"])


def load_blotter(path, findings):
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path, header=None, dtype=object)
    else:
        df = pd.read_excel(path, header=None)
    return _blotter_from_df(df, findings)


def load_blotter_bytes(data, filename, findings):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in EXCEL_EXTS:
        df = _table_from_excel_bytes(data)
    else:
        df = _table_from_text(data.decode("utf-8", errors="replace"))
    return _blotter_from_df(df, findings)


def load_blotter_text(text, findings):
    return _blotter_from_df(_table_from_text(text), findings)


def _accounting_number(value):
    """Parse '(21)' -> -21.0, '1,003.48' -> 1003.48."""
    s = str(value).strip().replace(",", "")
    if not s or s.lower() == "nan":
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        num = float(s)
    except ValueError:
        return 0.0
    return -num if neg else num


def _pcm_from_table(raw, findings):
    df = _named_table(raw, "Ticker")
    df = df[df["Ticker"].notna() & (df["Ticker"].astype(str).str.strip() != "")].copy()

    rows = []
    for _, r in df.iterrows():
        side = PCM_SIDE_MAP.get(str(r["Transaction Type"]).strip().lower())
        ticker = str(r["Ticker"]).strip().upper()
        if side is None:
            findings.add("ERROR", "PCM", "unknown_side", ticker,
                         detail=f"Unmapped PCM transaction type {r['Transaction Type']!r}")
            continue
        signed_qty = _accounting_number(r["Quantity"])
        if signed_qty == 0:
            continue
        expected_sign = SIDE_SIGN[side]
        if (signed_qty > 0) != (expected_sign > 0):
            findings.add("WARN", "PCM", "sign_mismatch", ticker, side,
                         right_value=str(signed_qty),
                         detail="PCM quantity sign disagrees with its own transaction type")
        rows.append({
            "ticker": ticker,
            "side": side,
            "qty": int(abs(signed_qty)),
            "price": _accounting_number(r["Price"]),
            "broker": str(r["Broker"]).strip().upper(),
            "fund": str(r["Fund"]).strip(),
        })

    return pd.DataFrame(rows, columns=["ticker", "side", "qty", "price", "broker", "fund"])


def load_pcm(path, findings):
    if str(path).lower().endswith(tuple(EXCEL_EXTS)):
        raw = pd.read_excel(path, header=None)
    else:
        raw = pd.read_csv(path, header=None, dtype=object, engine="python", on_bad_lines="skip")
    return _pcm_from_table(raw, findings)


def load_pcm_bytes(data, filename, findings):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in EXCEL_EXTS:
        raw = _table_from_excel_bytes(data)
    else:
        raw = _table_from_text(data.decode("utf-8", errors="replace"))
    return _pcm_from_table(raw, findings)


def load_pcm_text(text, findings):
    return _pcm_from_table(_table_from_text(text), findings)


def _tradefile_from_reader(reader, findings):
    orders = []
    allocs_by_order = {}
    trailer_count = None
    current = None

    for row in reader:
        if not row:
            continue
        rec = row[0].strip()
        if rec == "ORDER":
            current = {
                "sedol": row[9].strip().upper(),
                "side": TRADEFILE_SIDE_MAP.get(row[6]),
                "qty": int(row[18]) if row[18] else 0,
                "price": float(row[17]) if row[17] else None,
                "broker": row[14].strip().upper(),
                "order_id": row[2],
                "trade_date": row[3],
            }
            if current["side"] is None:
                findings.add("ERROR", "Tradefile", "unknown_side",
                             detail=f"Unmapped tradefile side {row[6]!r} on order {row[2]}")
            orders.append(current)
            allocs_by_order[current["order_id"]] = []
        elif rec == "ALLOCAT":
            if current is None:
                findings.add("ERROR", "Tradefile", "orphan_allocation", detail=f"ALLOCAT before any ORDER: {row}")
                continue
            allocs_by_order[current["order_id"]].append({
                "account": row[2],
                "account_name": TRADEFILE_ACCOUNT_NAMES.get(row[2], "unknown"),
                "qty": int(row[3]),
            })
        elif rec == "TRAILER":
            trailer_count = int(row[1])
        elif rec == "HEADER" or rec == "":
            continue
        else:
            findings.add("ERROR", "Tradefile", "unknown_record", detail=f"Unknown record type {rec!r}")

    # Internal integrity checks
    n_rows = len(orders) + sum(len(a) for a in allocs_by_order.values())
    if trailer_count is not None and trailer_count != n_rows:
        findings.add("ERROR", "Tradefile", "trailer_count",
                     left_value=str(trailer_count), right_value=str(n_rows),
                     detail="Trailer row count does not match parsed ORDER+ALLOCAT rows")

    for order in orders:
        alloc_sum = sum(a["qty"] for a in allocs_by_order[order["order_id"]])
        if alloc_sum != order["qty"]:
            findings.add("ERROR", "Tradefile", "allocation_sum", "", order["side"] or "",
                         left_value=f"order qty {order['qty']}",
                         right_value=f"alloc sum {alloc_sum}",
                         detail=f"Order {order['order_id']} (SEDOL {order['sedol']}) allocations do not sum to order quantity")
        unknown = [a["account"] for a in allocs_by_order[order["order_id"]] if a["account_name"] == "unknown"]
        if unknown:
            findings.add("WARN", "Tradefile", "unknown_account",
                         detail=f"Order {order['order_id']}: unrecognized allocation account(s) {unknown}")

    return pd.DataFrame(orders, columns=["sedol", "side", "qty", "price", "broker", "order_id", "trade_date"])


def load_tradefile(path, findings):
    with open(path, newline="") as f:
        return _tradefile_from_reader(csv.reader(f, delimiter="|"), findings)


def load_tradefile_bytes(data, filename, findings):
    return load_tradefile_text(data.decode("utf-8", errors="replace"), findings)


def load_tradefile_text(text, findings):
    return _tradefile_from_reader(csv.reader(StringIO(text), delimiter="|"), findings)


# ---------------------------------------------------------------------------
# SEDOL -> ticker resolution
# ---------------------------------------------------------------------------

def resolve_sedols(sedols, emsx_df, findings):
    """Map SEDOLs to tickers: EMSX first, then Backtester security master, then xbbg."""
    mapping = dict(zip(emsx_df["sedol"], emsx_df["ticker"]))
    missing = [s for s in sedols if s not in mapping]

    if missing:
        sm_map = _resolve_via_security_master(missing)
        mapping.update(sm_map)
        missing = [s for s in missing if s not in sm_map]
        if sm_map:
            print(f"Resolved {len(sm_map)} SEDOL(s) via Backtester security master.")

    if missing:
        bbg_map = _resolve_via_xbbg(missing)
        mapping.update(bbg_map)
        missing = [s for s in missing if s not in bbg_map]
        if bbg_map:
            print(f"Resolved {len(bbg_map)} SEDOL(s) via Bloomberg (xbbg).")

    for sedol in missing:
        findings.add("ERROR", "Tradefile", "unresolved_sedol",
                     detail=f"SEDOL {sedol} not found in EMSX, security master, or Bloomberg")
    return mapping


def _resolve_via_security_master(sedols):
    """Backtester SecurityMaster: local parquet, no network. Its config loads
    relative to the repo root, so chdir there for the duration of the import."""
    prev_cwd = os.getcwd()
    try:
        sys.path.insert(0, BACKTESTER_REPO)
        os.chdir(BACKTESTER_REPO)
        import backtester as bt
        sm = bt.SecurityMaster()
        out = {}
        for sedol in sedols:
            ticker = sm.map_id(sedol, "bloomberg_ticker")
            if ticker:
                out[sedol] = str(ticker).upper()
        return out
    except Exception as e:
        print(f"Security master unavailable ({type(e).__name__}: {e}); skipping.")
        return {}
    finally:
        os.chdir(prev_cwd)
        if BACKTESTER_REPO in sys.path:
            sys.path.remove(BACKTESTER_REPO)


def _resolve_via_xbbg(sedols):
    """Bloomberg terminal lookup via xbbg, if installed and terminal is running."""
    try:
        from xbbg import blp
        ids = [f"/sedol/{s}" for s in sedols]
        res = blp.bdp(ids, "TICKER")
        out = {}
        for bbg_id, row in res.iterrows():
            sedol = str(bbg_id).replace("/sedol/", "").split()[0].upper()
            ticker = row.get("ticker") or row.get("TICKER")
            if pd.notna(ticker):
                out[sedol] = str(ticker).upper()
        return out
    except Exception as e:
        print(f"xbbg unavailable ({type(e).__name__}: {e}); skipping.")
        return {}


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

def aggregate(df):
    """Per (ticker, side): total qty and quantity-weighted average price."""
    if df.empty:
        return pd.DataFrame(columns=["ticker", "side", "qty", "wavg_price"])
    df = df.copy()
    df["_pxqty"] = df["price"].fillna(0) * df["qty"]
    g = df.groupby(["ticker", "side"], as_index=False).agg(qty=("qty", "sum"), _pxqty=("_pxqty", "sum"))
    g["wavg_price"] = g["_pxqty"] / g["qty"]
    return g.drop(columns="_pxqty")


def net_by_ticker(df):
    if df.empty:
        return pd.Series(dtype=int)
    signed = df["qty"] * df["side"].map(SIDE_SIGN)
    return signed.groupby(df["ticker"]).sum().astype(int)


def compare_pair(left_agg, right_agg, left_name, right_name, qty_tol, price_tol):
    """Bidirectional comparison of two aggregated frames on (ticker, side).

    Returns a list of discrepancy dicts. Each carries ``net_ok``: True when the
    net signed quantity per ticker still agrees between the two sources, which
    means the disagreement is a side-classification difference that step 4 may
    resolve against the Blotter EOD. Discrepancies that change net quantity are
    real missing/extra trades and always stay errors.
    """
    merged = left_agg.merge(right_agg, on=["ticker", "side"], how="outer",
                            suffixes=("_l", "_r"), indicator=True)

    left_net = net_by_ticker_from_agg(left_agg)
    right_net = net_by_ticker_from_agg(right_agg)

    discrepancies = []
    net_error_tickers = set()

    for _, r in merged.iterrows():
        ticker, side = r["ticker"], r["side"]
        net_ok = left_net.get(ticker, 0) == right_net.get(ticker, 0)

        if r["_merge"] == "left_only":
            if not net_ok:
                net_error_tickers.add(ticker)
            discrepancies.append({
                "check": "missing_bucket", "ticker": ticker, "side": side, "net_ok": net_ok,
                "left_value": f"qty {int(r['qty_l'])}", "right_value": "absent",
                "left_qty": int(r["qty_l"]), "right_qty": None,
                "detail": f"In {left_name} but not in {right_name}",
            })
        elif r["_merge"] == "right_only":
            if not net_ok:
                net_error_tickers.add(ticker)
            discrepancies.append({
                "check": "extra_bucket", "ticker": ticker, "side": side, "net_ok": net_ok,
                "left_value": "absent", "right_value": f"qty {int(r['qty_r'])}",
                "left_qty": None, "right_qty": int(r["qty_r"]),
                "detail": f"In {right_name} but not in {left_name}",
            })
        else:
            qty_diff = int(r["qty_r"]) - int(r["qty_l"])
            if abs(qty_diff) > qty_tol:
                if not net_ok:
                    net_error_tickers.add(ticker)
                discrepancies.append({
                    "check": "qty_mismatch", "ticker": ticker, "side": side, "net_ok": net_ok,
                    "left_value": str(int(r["qty_l"])), "right_value": str(int(r["qty_r"])),
                    "left_qty": int(r["qty_l"]), "right_qty": int(r["qty_r"]),
                    "detail": f"Quantity differs by {qty_diff:+d} ({right_name} minus {left_name})",
                })
            px_l, px_r = r["wavg_price_l"], r["wavg_price_r"]
            if pd.notna(px_l) and pd.notna(px_r) and px_r != 0 and px_l != 0:
                if abs(px_r - px_l) > price_tol and abs(px_r - px_l) / px_l > 0.0005:
                    discrepancies.append({
                        "check": "price_mismatch", "ticker": ticker, "side": side, "net_ok": True,
                        "left_value": f"{px_l:.4f}", "right_value": f"{px_r:.4f}",
                        "left_qty": None, "right_qty": None, "price_only": True,
                        "detail": f"Weighted-avg price differs by {px_r - px_l:+.4f}",
                    })

    # One underlying discrepancy is counted once: the net-level check only fires
    # for tickers with no bucket-level error already recorded above.
    for ticker in sorted(set(left_net.index) | set(right_net.index)):
        l, r = left_net.get(ticker, 0), right_net.get(ticker, 0)
        if l != r and ticker not in net_error_tickers:
            discrepancies.append({
                "check": "net_qty_mismatch", "ticker": ticker, "side": "", "net_ok": False,
                "left_value": str(l), "right_value": str(r),
                "left_qty": None, "right_qty": None,
                "detail": f"Net signed quantity differs by {r - l:+d} ({right_name} minus {left_name})",
            })

    return discrepancies


def _ticker_side_maps(agg):
    """{ticker: {side: qty}} from an aggregated frame."""
    maps = {}
    for _, r in agg.iterrows():
        maps.setdefault(r["ticker"], {})[r["side"]] = int(r["qty"])
    return maps


def _maps_match(a, b, qty_tol):
    return a is not None and b is not None and set(a) == set(b) and all(
        abs(a[side] - b[side]) <= qty_tol for side in a
    )


def adjudicate_with_blotter(discrepancies, left_agg, right_agg, blotter_agg,
                            left_name, right_name, step, findings, qty_tol):
    """Step 4: settle side-classification discrepancies against the Blotter EOD.

    A discrepancy is resolvable only when net quantity per ticker still matches
    (pure side/classification disagreement). The blotter's per-ticker bucket
    profile (side -> qty) is compared against each source's profile; when it
    matches one source exactly, that source is corroborated and every net-ok
    discrepancy for the ticker becomes RESOLVED as a single classification
    difference. Real missing/extra trades (net quantity changed) always remain
    errors, as do discrepancies the blotter cannot settle.
    """
    left_maps = _ticker_side_maps(left_agg)
    right_maps = _ticker_side_maps(right_agg)
    blotter_maps = _ticker_side_maps(blotter_agg)

    source = f"{left_name} vs {right_name}"
    for d in discrepancies:
        if d.get("price_only"):
            findings.add("WARN", source, d["check"], d["ticker"], d["side"],
                         left_value=d["left_value"], right_value=d["right_value"],
                         detail=d["detail"], step=step)
            continue

        if not d["net_ok"]:
            findings.add("ERROR", source, d["check"], d["ticker"], d["side"],
                         left_value=d["left_value"], right_value=d["right_value"],
                         detail=d["detail"], step=step)
            continue

        ticker = d["ticker"]
        blotter_map = blotter_maps.get(ticker)
        if _maps_match(blotter_map, left_maps.get(ticker), qty_tol):
            supported = left_name
        elif _maps_match(blotter_map, right_maps.get(ticker), qty_tol):
            supported = right_name
        else:
            supported = None

        if supported:
            shown = ", ".join(f"{s} {q}" for s, q in sorted(blotter_map.items()))
            findings.add("RESOLVED", source, d["check"], ticker, d["side"],
                         left_value=d["left_value"], right_value=d["right_value"],
                         detail=d["detail"] + f". Resolved: Blotter EOD shows {ticker} "
                                f"({shown}), corroborating {supported}; "
                                "side classification difference.",
                         step=step)
        else:
            findings.add("ERROR", source, d["check"], ticker, d["side"],
                         left_value=d["left_value"], right_value=d["right_value"],
                         detail=d["detail"] + ". Net quantity matches but the Blotter EOD "
                                "does not corroborate either source.",
                         step=step)


def net_by_ticker_from_agg(agg):
    if agg.empty:
        return pd.Series(dtype=int)
    signed = agg["qty"] * agg["side"].map(SIDE_SIGN)
    return signed.groupby(agg["ticker"]).sum().astype(int)


def check_crosses(cross_df, source_name, findings):
    """Internal fund-to-fund crosses must net to zero per ticker."""
    if cross_df.empty:
        return
    net = net_by_ticker(cross_df)
    for ticker, n in net.items():
        if n != 0:
            findings.add("ERROR", source_name, "cross_not_flat", ticker,
                         left_value="0", right_value=str(int(n)),
                         detail="Internal cross (CROS) legs do not net to zero",
                         step="Crosses")


def compare_crosses(blotter_x, pcm_x, findings):
    """Cross volume per ticker should agree between the blotter and PCM."""
    def gross_buys(df):
        if df.empty:
            return pd.Series(dtype=int)
        buys = df[df["side"].map(SIDE_SIGN) > 0]
        return buys.groupby("ticker")["qty"].sum().astype(int)

    b, p = gross_buys(blotter_x), gross_buys(pcm_x)
    for ticker in sorted(set(b.index) | set(p.index)):
        bq, pq = int(b.get(ticker, 0)), int(p.get(ticker, 0))
        if bq != pq:
            findings.add("ERROR", "PCM", "cross_qty_mismatch", ticker,
                         left_value=f"blotter cross {bq}", right_value=f"pcm cross {pq}",
                         detail="Internal cross volume differs between blotter and PCM",
                         step="Crosses")


# ---------------------------------------------------------------------------
# Reconciliation (shared by CLI and web app)
# ---------------------------------------------------------------------------

def reconcile(emsx, blotter, pcm, tradefile, findings, qty_tol=0, price_tol=0.01,
              goldman_brokers=DEFAULT_GOLDMAN_BROKERS,
              pcm_excluded_emsx_brokers=EMSX_PCM_EXCLUDED_BROKERS):
    """Run the four-step workflow on already-loaded DataFrames; returns the report.

    Step 1: EMSX (brokers GSPT/PREX/GSOP) vs Goldman tradefile, bidirectional.
    Step 2: EMSX (excluding TDAI)         vs PCM, bidirectional.
    Step 3: PCM  (brokers GSPT/PREX/GSOP) vs Goldman tradefile, bidirectional.
    Step 4: Blotter EOD adjudication of the discrepancies from steps 1-3.
    """
    goldman_brokers = {str(b).strip().upper() for b in goldman_brokers}
    pcm_excluded = {str(b).strip().upper() for b in pcm_excluded_emsx_brokers}

    sedol_map = resolve_sedols(tradefile["sedol"].unique().tolist(), emsx, findings)
    tradefile = tradefile.assign(ticker=tradefile["sedol"].map(sedol_map))
    tradefile = tradefile[tradefile["ticker"].notna()]

    # Internal crosses (CROS) never route through EMSX: split them out, check
    # they net flat, and compare cross volume between blotter and PCM.
    blotter_x = blotter[blotter["broker"] == CROSS_BROKER]
    pcm_x = pcm[pcm["broker"] == CROSS_BROKER]
    blotter = blotter[blotter["broker"] != CROSS_BROKER]
    pcm = pcm[pcm["broker"] != CROSS_BROKER]
    check_crosses(blotter_x, "Blotter", findings)
    check_crosses(pcm_x, "PCM", findings)
    compare_crosses(blotter_x, pcm_x, findings)

    tradefile_agg = aggregate(tradefile)
    pcm_agg = aggregate(pcm)

    # Step 1: EMSX Goldman scope vs tradefile.
    emsx_goldman_agg = aggregate(emsx[emsx["broker"].isin(goldman_brokers)])
    step1 = compare_pair(emsx_goldman_agg, tradefile_agg, "EMSX", "Tradefile",
                         qty_tol, price_tol)

    # Step 2: EMSX without TDAI vs PCM.
    emsx_no_tdai_agg = aggregate(emsx[~emsx["broker"].isin(pcm_excluded)])
    step2 = compare_pair(emsx_no_tdai_agg, pcm_agg, "EMSX", "PCM",
                         qty_tol, price_tol)

    # Step 3: Goldman-scope PCM vs tradefile.
    pcm_goldman_agg = aggregate(pcm[pcm["broker"].isin(goldman_brokers)])
    step3 = compare_pair(pcm_goldman_agg, tradefile_agg, "PCM", "Tradefile",
                         qty_tol, price_tol)

    # Step 4: adjudicate discrepancies against the Blotter EOD, scoped like the
    # comparison that produced them.
    blotter_goldman_agg = aggregate(blotter[blotter["broker"].isin(goldman_brokers)])
    blotter_no_tdai_agg = aggregate(blotter[~blotter["broker"].isin(pcm_excluded)])
    adjudicate_with_blotter(step1, emsx_goldman_agg, tradefile_agg, blotter_goldman_agg,
                            "EMSX", "Tradefile", "Step 1", findings, qty_tol)
    adjudicate_with_blotter(step2, emsx_no_tdai_agg, pcm_agg, blotter_no_tdai_agg,
                            "EMSX", "PCM", "Step 2", findings, qty_tol)
    adjudicate_with_blotter(step3, pcm_goldman_agg, tradefile_agg, blotter_goldman_agg,
                            "PCM", "Tradefile", "Step 3", findings, qty_tol)

    return findings.to_frame()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def run(emsx_path, blotter_path, pcm_path, tradefile_path, out_path, qty_tol, price_tol,
        goldman_brokers=DEFAULT_GOLDMAN_BROKERS):
    findings = Findings()

    print(f"Loading EMSX: {emsx_path}")
    emsx = load_emsx(emsx_path, findings)
    print(f"  {len(emsx)} filled EMSX rows, {emsx['ticker'].nunique()} tickers")

    print(f"Loading Blotter EOD: {blotter_path}")
    blotter = load_blotter(blotter_path, findings)
    print(f"  {len(blotter)} filled blotter trades")

    print(f"Loading PCMBlotter: {pcm_path}")
    pcm = load_pcm(pcm_path, findings)
    print(f"  {len(pcm)} PCM trades")

    print(f"Loading tradefile: {tradefile_path}")
    tradefile = load_tradefile(tradefile_path, findings)
    print(f"  {len(tradefile)} tradefile orders")

    report = reconcile(emsx, blotter, pcm, tradefile, findings,
                       qty_tol=qty_tol, price_tol=price_tol, goldman_brokers=goldman_brokers)
    report.to_csv(out_path, index=False)

    print("\n" + "=" * 70)
    n_err = (report["severity"] == "ERROR").sum()
    n_warn = (report["severity"] == "WARN").sum()
    n_resolved = (report["severity"] == "RESOLVED").sum()
    print(f"RESULT: {n_err} error(s), {n_warn} warning(s), "
          f"{n_resolved} resolved by blotter  ->  {out_path}")
    print("=" * 70)
    if not report.empty:
        with pd.option_context("display.max_rows", None, "display.max_colwidth", 90, "display.width", 250):
            print(report.to_string(index=False))
    else:
        print("All comparisons reconcile. No differences found.")

    return report


def main():
    # Windows consoles default to cp1252; keep report text printable regardless.
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    p = argparse.ArgumentParser(description="Reconcile EOD trade files against the EMSX export.")
    p.add_argument("--emsx", required=True, help="EMSX grid export (.xlsx)")
    p.add_argument("--blotter", required=True, help="Vantage BlotterEOD (.xlsx or .csv)")
    p.add_argument("--pcm", required=True, help="PCMBlotter (.csv)")
    p.add_argument("--tradefile", required=True, help="Goldman tradefile.nine82 (.csv, pipe-delimited)")
    p.add_argument("--out", default=None, help="Findings CSV path (default: timestamped in cwd)")
    p.add_argument("--qty-tol", type=int, default=0, help="Allowed share difference before flagging (default 0)")
    p.add_argument("--price-tol", type=float, default=0.01, help="Allowed absolute price difference (default 0.01)")
    p.add_argument("--goldman-brokers", default=",".join(sorted(DEFAULT_GOLDMAN_BROKERS)),
                   help="Comma-separated Goldman broker codes for the tradefile/PCM Goldman scopes")
    args = p.parse_args()

    out = args.out or f"blottercheck_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    goldman = {b.strip().upper() for b in args.goldman_brokers.split(",") if b.strip()}
    report = run(args.emsx, args.blotter, args.pcm, args.tradefile, out, args.qty_tol, args.price_tol,
                 goldman_brokers=goldman)
    sys.exit(1 if (report["severity"] == "ERROR").any() else 0)


if __name__ == "__main__":
    main()
