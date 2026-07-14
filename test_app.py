"""Exercise the /check endpoint with the sample files, both as uploads and pastes."""
import io
import sys

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from eod import app

client = app.test_client()


def post(data):
    resp = client.post("/check", data=data, content_type="multipart/form-data")
    body = resp.get_json()
    assert body is not None, resp.data[:500]
    return resp.status_code, body


def file_part(path):
    return (open(path, "rb"), path.split("/")[-1])


# --- Case 1: all four as file uploads -------------------------------------
code, body = post({
    "emsx_file": file_part("samples/grid.xlsx"),
    "blotter_file": file_part("samples/BlotterEOD07.10.26.xlsx"),
    "pcm_file": file_part("samples/PCMBlotter 2026-07-10.csv"),
    "tradefile_file": file_part("samples/tradefile.nine82.20260710045714.csv"),
})
print("uploads:", code, body["ok"], body.get("summary"))
assert code == 200 and body["ok"] and body["summary"]["errors"] == 0, body

# --- Case 2: all four as pasted text ---------------------------------------
# EMSX and blotter pasted as tab-separated grids (like copying cells in Excel).
wb = load_workbook("samples/grid.xlsx", read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
emsx_text = "\n".join(
    "\t".join("" if c is None else str(c) for c in row)
    for row in ws.iter_rows(values_only=True)
)
wb.close()

blotter_df = pd.read_excel("samples/BlotterEOD07.10.26.xlsx", header=None)
blotter_text = blotter_df.to_csv(sep="\t", index=False, header=False)

pcm_text = open("samples/PCMBlotter 2026-07-10.csv", encoding="utf-8", errors="replace").read()
tradefile_text = open("samples/tradefile.nine82.20260710045714.csv").read()

code, body = post({
    "emsx_text": emsx_text,
    "blotter_text": blotter_text,
    "pcm_text": pcm_text,
    "tradefile_text": tradefile_text,
})
print("pastes:", code, body["ok"], body.get("summary"), body.get("error", ""))
assert code == 200 and body["ok"] and body["summary"]["errors"] == 0, body

# --- Case 3: tampered paste shows diffs -------------------------------------
tampered = tradefile_text.replace("|156.4000|6|", "|156.4000|9|")  # qty 6 -> 9
code, body = post({
    "emsx_text": emsx_text,
    "blotter_text": blotter_text,
    "pcm_text": pcm_text,
    "tradefile_text": tampered,
})
print("tampered:", code, body["ok"], body.get("summary"))
findings = body["findings"]
checks = {(f["check"], f["ticker"]) for f in findings}
print("findings:", checks)
assert body["summary"]["errors"] >= 2, body["summary"]  # alloc sum + qty/net mismatch

# --- Case 4: missing source rejected ----------------------------------------
code, body = post({"emsx_text": emsx_text})
print("missing source:", code, body.get("error"))
assert code == 400 and not body["ok"]

print("\nALL WEB TESTS PASSED")
